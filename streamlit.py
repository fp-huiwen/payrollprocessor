### define imports ###
import streamlit as st

import pandas as pd
import numpy as np
import os
import zipfile

# handling merged columns
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.dataframe import dataframe_to_rows

# handling date mapping
from datetime import datetime
from dateutil.relativedelta import relativedelta
from calendar import monthrange

# handling password protected files
import io
import msoffcrypto

### define functions and variables ###

def read_workbook(business_unit, uploaded_file, psword, sheet):
    decrypted_workbook = io.BytesIO()

    # Check if the file is password protected
    if psword == "":
        decrypted_workbook = uploaded_file  # directly use the uploaded file, which is already a BytesIO object
    else: 
        # Decrypt the file if a password is provided
        office_file = msoffcrypto.OfficeFile(uploaded_file)
        office_file.load_key(password=psword)
        office_file.decrypt(decrypted_workbook)

    # Check the business unit and proceed with Genie or other types
    if business_unit == "Genie":
        wb = load_workbook(decrypted_workbook, data_only=True)  # Load with openpyxl
        ws = wb.active  # Get the active worksheet (or specify sheet name)

        # Unmerge cells in Genie-specific sheets
        mcr_coord_list = [mcr.coord for mcr in ws.merged_cells.ranges]
        for mcr in mcr_coord_list:
            min_col, min_row, max_col, max_row = range_boundaries(mcr)
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(mcr)
            for row in ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value

        # Convert worksheet data to DataFrame
        data = ws.iter_rows(values_only=True)
        df = pd.DataFrame(data)

        # Process column and row headers
        df.iloc[:, 0] = pd.to_numeric(df.iloc[:, 0], errors='coerce')
        first_col = df[df.iloc[:, 0] >= 0].index
        column_row = first_col[0]

        # Handling unmerged cells (potential duplicate column names)
        col_row_1 = [cell.value for cell in ws[int(column_row - 1)]]
        col_row_2 = [cell.value for cell in ws[int(column_row)]]

        final_columns = []
        for col1, col2 in zip(col_row_1, col_row_2):
            if col1 == col2 or col1 is None:
                final_columns.append(col2)
            else:
                final_columns.append(f"{col1}_{col2}")

        # Extract data from the relevant rows
        data = ws.iter_rows(min_row=column_row + 1, values_only=True)
        df = pd.DataFrame(data, columns=final_columns)

    else:
        # For non-Genie files, handle Excel decryption and sheet reading
        df = pd.read_excel(decrypted_workbook, sheet_name=sheet)
        df = df.dropna(axis=1, how="all")  # Remove empty columns

        # Find the first row with numeric data in the first column
        df.iloc[:, 0] = pd.to_numeric(df.iloc[:, 0], errors='coerce')
        first_col = df[df.iloc[:, 0] >= 0].index

        if not first_col.empty:
            header_index = first_col[0]
            df = pd.read_excel(decrypted_workbook, sheet_name=sheet, skiprows=header_index)
            df = df.drop(df.columns[[0]], axis=1)  # Drop the first column (empty)
        else:
            raise ValueError("Check data type of the first column")

    # Remove rows with non-numerical first columns
    df = df[pd.to_numeric(df.iloc[:, 0], errors='coerce').notna()]

    # Clean column names by removing line breaks
    df.columns = df.columns.str.split('\n').str[0]

    return df

renaming_dict = {
    "Lelang Sion": {
        "employee_id": "ID Personalia",
        "department": "Bagian",
        # check join and end date
        "hire_date": "Join Date",
        "termination_date": "Resign Date",
        "basic_salary": ["Gaji Pokok (Basic Salary)", "Prorate"],
        "gross_salary": "Gross Income",
        # check need to split?
        "employer_statutory_payments": ["Tunjangan Premi JKK BPJS Ketenagakerjaan (P)", "Tunjangan Premi JKM BPJS Ketenagakerjaan (P)", "Tunjangan Premi JHT BPJS Ketenagakerjaan (P)", "Tunjangan Premi JP BPJS Ketenagakerjaan (P)", "Premi BPJS Kesehatan (P)", "Tunjangan PPh21", "Tunjangan PPh21 Insentif/Bonus"],
        "overtime": ["Lembur"],
        "claims": 0,
        "allowance": ["Tunjangan Acting", "Attendance Allowance", "Communication Allowance", "Health Allowance", "Meal Allowance", "Position Allowance", "Skill Allowance", "Transport Allowance", "Tunjangan Covid", "Tunjangan Jabatan", "Tunjangan Klasifikasi", "Other Allowance"],
        "bonus": "Bonus",
        "incentives": ["Insentif"],
        "others": 0
    },
    "Jualo": {
        "employee_id": "ID Personalia",
        "department": "Bagian",
        # check join and end date
        "hire_date": "Join Date",
        "termination_date": "Resign Date",
        "basic_salary": ["Gaji Pokok (Basic Salary)", "Prorate"],
        "gross_salary": "Gross Income",
        # check need to split?
        "employer_statutory_payments": ["Tunjangan Premi JKK BPJS Ketenagakerjaan (P)", "Tunjangan Premi JKM BPJS Ketenagakerjaan (P)", "Tunjangan Premi JHT BPJS Ketenagakerjaan (P)", "Tunjangan Premi JP BPJS Ketenagakerjaan (P)", "Premi BPJS Kesehatan (P)", "Tunjangan PPh21", "Tunjangan PPh21 Insentif/Bonus"],
        "overtime": ["Lembur"],
        "claims": 0,
        "allowance": ["Tunjangan Acting", "Attendance Allowance", "Communication Allowance", "Health Allowance", "Meal Allowance", "Position Allowance", "Skill Allowance", "Transport Allowance", "Tunjangan Covid", "Tunjangan Jabatan", "Tunjangan Klasifikasi", "Other Allowance"],
        "bonus": "Bonus",
        "incentives": ["Insentif"],
        "others": 0
    },
    "PMN": {
        "employee_id": "ID Personalia",
        "department": "Bagian",
        # check join and end date
        "hire_date": "Join Date",
        "termination_date": "Resign Date",
        "basic_salary": ["Gaji Pokok (Basic Salary)", "Prorate"],
        "gross_salary": "Gross Income",
        # check need to split?
        "employer_statutory_payments": ["Tunjangan Premi JKK BPJS Ketenagakerjaan (P)", "Tunjangan Premi JKM BPJS Ketenagakerjaan (P)", "Tunjangan Premi JHT BPJS Ketenagakerjaan (P)", "Tunjangan Premi JP BPJS Ketenagakerjaan (P)", "Premi BPJS Kesehatan (P)", "Tunjangan PPh21 Reguler", "Tunjangan PPh 21 Insentif/Bonus"],
        "overtime": ["Lembur"],
        "claims": 0,
        "allowance": ["Tunjangan Acting", "Attendance Allowance", "Communication Allowance", "Health Allowance", "Meal Allowance", "Position Allowance", "Skill Allowance", "Transport Allowance", "Tunjangan Covid", "Tunjangan Jabatan", "Tunjangan Klasifikasi", "Other Allowance"],
        "bonus": "Bonus",
        "incentives": ["Insentif"],
        "others": 0
    },
    "Carro": {
        "employee_id": "ID Personalia",
        "department": "Bagian",
        "hire_date": "Join Date",
        "termination_date": "Resign Date",
        # check join and end date
        "basic_salary": ["Gaji Pokok (Basic Salary)", "Prorate"],
        "gross_salary": "Gross Income",
        # check need to split?
        "employer_statutory_payments": ["Tunjangan Premi JKK BPJS Ketenagakerjaan (P)", "Tunjangan Premi JKM BPJS Ketenagakerjaan (P)", "Tunjangan Premi JHT BPJS Ketenagakerjaan (P)", "Tunjangan Premi JP BPJS Ketenagakerjaan (P)", "Premi BPJS Kesehatan (P)", "Tunjangan PPh21 Reguler", "Tunjangan PPh 21 Insentif/Bonus"],
        "overtime": ["Lembur"],
        "claims": 0,
        "allowance": ["Tunjangan Acting", "Attendance Allowance", "Communication Allowance", "Health Allowance", "Meal Allowance", "Position Allowance", "Skill Allowance", "Transport Allowance", "Tunjangan Covid", "Tunjangan Jabatan", "Tunjangan Klasifikasi", "Other Allowance"],
        "bonus": "Bonus",
        "incentives": ["Insentif"],
        "others": 0
    },
    "WMN": {
        "employee_id": "ID Personalia",
        "department": "Bagian",
        "hire_date": "Join Date",
        "termination_date": "Resign Date",
        # check join and end date
        "basic_salary": ["Gaji Pokok (Basic Salary)", "Prorate"],
        "gross_salary": "Gross Income",
        # check need to split?
        "employer_statutory_payments": ["Tunjangan Premi JKK BPJS Ketenagakerjaan (P)", "Tunjangan Premi JKM BPJS Ketenagakerjaan (P)", "Tunjangan Premi JHT BPJS Ketenagakerjaan (P)", "Tunjangan Premi JP BPJS Ketenagakerjaan (P)", "Premi BPJS Kesehatan (P)", "Tunjangan PPh21", "Tunjangan PPh21 Insentif/Bonus"],
        "overtime": ["Lembur"],
        "claims": 0,
        "allowance": ["Tunjangan Acting", "Attendance Allowance", "Communication Allowance", "Health Allowance", "Meal Allowance", "Position Allowance", "Skill Allowance", "Transport Allowance", "Tunjangan Covid", "Tunjangan Jabatan", "Tunjangan Klasifikasi", "Other Allowance"],
        "bonus": "Bonus",
        "incentives": ["Insentif"],
        "others": 0
    },
    "KMN": {
        "employee_id": "ID Personalia",
        "department": "Bagian",
        "hire_date": "Join Date",
        "termination_date": "Resign Date",
        # check join and end date
        "basic_salary": ["Gaji Pokok (Basic Salary)", "Prorate"],
        "gross_salary": "Gross Income",
        # check need to split?
        "employer_statutory_payments": ["Tunjangan Premi JKK BPJS Ketenagakerjaan (P)", "Tunjangan Premi JKM BPJS Ketenagakerjaan (P)", "Tunjangan Premi JHT BPJS Ketenagakerjaan (P)", "Tunjangan Premi JP BPJS Ketenagakerjaan (P)", "Premi BPJS Kesehatan (P)", "Tunjangan PPh21", "Tunjangan PPh21 Insentif/Bonus"],
        "overtime": ["Lembur"],
        "claims": 0,
        "allowance": ["Acting Allowance", "Attendance Allowance", "Communication Allowance", "Health Allowance", "Meal Allowance", "Position Allowance", "Skill Allowance", "Transport Allowance", "Tunjangan Covid", "Tunjangan Jabatan", "Tunjangan Klasifikasi", "Other Allowance"],
        "bonus": "Bonus",
        "incentives": ["Insentif"],
        "others": 0
    },
    "Genie": {
        # end date added if employee left (might not be in every month)
        "employee_id": "Employee ID",
        "department": "Organization",
        # check join and end date
        "hire_date": "Join date",
        "termination_date": 0,
        "basic_salary": "Basic Salary",
        "gross_salary": "Gross",
        # check need to split?
        "employer_statutory_payments": ['PPH 21 Benefit', 'JKK', 'JKM', 'JHT Company', 'JP Company', 'BPJS K Company'],
        "overtime": "Allowance_Overtime",
        "claims": 0,
        "allowance": ['Allowance_Position', 'Allowance_Transportation', 'Allowance_Communication', 'Allowance_Survey', 'Allowance_Other', 'Allowance_BPJSK Employee', 'Allowance_JHT Employee', 'Allowance_JP Employee', 'Allowance_Rapel'],
        "bonus": 0,
        "incentives": 0,
        "others": 0
    }
}

def format_columns(month, business_unit, df):

    def clean_numeric_columns(column):
        numeric_column = pd.to_numeric(column, errors = "coerce")
        return numeric_column

    dct = renaming_dict.get(business_unit)
    columns = list(dct.keys())

    # add/rename relevant columns
    for new_column in columns:
        old_column = dct.get(new_column)
        if isinstance(old_column, str):
            df.rename(columns = {old_column: new_column}, inplace = True)

            # format date columns
            if "date" in str.lower(old_column):
                df[new_column] = df[new_column].astype(str)
                # changes invalid entries into NAs
                df[new_column] = df[new_column].apply(lambda x: pd.NaT if len(x) == 1 else x)
                df[new_column] = pd.to_datetime(df[new_column], errors = "coerce")
        
        # column doesn't exist (output 0)
        elif isinstance(old_column, int):
            # fills missing date columns with NAs
            if "date" in str.lower(new_column):
                df[new_column] = np.nan
            else:
                df[new_column] = 0

        # list (columns should be numerical)
        else:
            # handle potential inconsistent formatting
            df[old_column] = df[old_column].apply(clean_numeric_columns)
            df[new_column] = df[old_column].sum(axis = 1)

    # add boolean columns
    current_year = datetime.now().year
    month = datetime.strptime(month, "%B").month 
    current_date = datetime(current_year, month, 1)
    start_filter = (current_date - relativedelta(months = 1)).replace(day = 21)
    last_day = monthrange(current_year, month)[1]
    end_filter = datetime(current_year, month, last_day)
    
    # temporarily handles missing date function
    if df["hire_date"].isna().all():
        df["is_new"] = 0
    else:
        df["is_new"] = (df["hire_date"] >= start_filter).astype(int)
        
    if df["termination_date"].isna().all():
        df["is_resign"] = 0
    else:
        df["is_resign"] = (df["termination_date"] <= end_filter).astype(int)

    # add total staff cost
    # [ID] gross salary include employer statutory payments (equivalent to total staff cost)
    df["total_staff_cost"] = df["gross_salary"]
    df["gross_salary"] = df["gross_salary"] - df["employer_statutory_payments"]

    columns.extend(["is_new", "is_resign", "total_staff_cost"])

    # soft formatting of column names
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.replace(':', '')
    df.columns = df.columns.str.replace('/', '')
    df.columns = df.columns.str.replace('-', '_')
    df.columns = df.columns.str.replace(r'\s+', ' ', regex = True)
    df.columns = df.columns.str.replace(' ', '_')
    df.columns = df.columns.str.lower()

    # remove multiple underscores and reduce to a single underscore
    df.columns = df.columns.str.replace(r'_+', '_', regex=True)

    return df

def add_df_to_xlsm(xlsm_file, df, new_sheet_name):
    try:
        # Load the existing workbook, keeping VBA intact
        book = load_workbook(xlsm_file, keep_vba=True)

        # Create a new sheet or get the existing one
        if new_sheet_name in book.sheetnames:
            sheet = book[new_sheet_name]
            sheet.delete_rows(1, sheet.max_row)  # Clear existing content
        else:
            sheet = book.create_sheet(new_sheet_name)

        # Write the DataFrame to the sheet
        for r in dataframe_to_rows(df, index=False, header=True):
            sheet.append(r)

        # Save the workbook to an in-memory buffer
        output = io.BytesIO()
        book.save(output)
        output.seek(0)
        
        return output

    except Exception as e:
        st.error(f"An error occurred: {str(e)}")
        return None

### website ###

st.title("Payroll Summary Processor")

# Define the list of months
months = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]

# Add a dropdown (selectbox) for users to select the month
selected_month = st.selectbox("Select Month", months)

# Initialize session state to store uploaded files and business units
if 'uploaded_files' not in st.session_state:
    st.session_state.uploaded_files = []
if 'business_units' not in st.session_state:
    st.session_state.business_units = []
if 'passwords' not in st.session_state:
    st.session_state.passwords = []
if 'sheets' not in st.session_state:
    st.session_state.sheets = []

# Define the list of business units
business_units = ["Lelang Sion", "Jualo", "PMN", "Carro", "WMN", "KMN", "Genie"]

# Create a file uploader and business unit selector for each file upload
st.write("Upload a payroll file and select the corresponding business unit (enter the password and sheet name if necessary)")

# File uploader
uploaded_file = st.file_uploader("Upload Monthly Payroll File", type = ['xlsx'], key = "key1")

# Business unit selector
selected_business_unit = st.selectbox("Select Business Unit for the Uploaded File", business_units)

# Option to enter a password for password-protected files
password = st.text_input("Enter password (leave blank if none):", type="password")

# Option to enter sheet name (if needed)
sheet_name = st.text_input("Enter sheet name (optional, defaults to first sheet):", value="")

# Add a button to confirm the file and business unit selection
if uploaded_file:
    if st.button("Add File"):
        # Store the uploaded file and the selected business unit in session state
        st.session_state.uploaded_files.append(uploaded_file)
        st.session_state.business_units.append(selected_business_unit)
        st.session_state.passwords.append(password)
        st.session_state.sheets.append(sheet_name)
        st.success(f"File {uploaded_file.name} for {selected_business_unit} added successfully!")
        
st.write("Upload the existing macros-enabled Payroll Summary Workbook")
        
uploaded_xlsm_file = st.file_uploader("Upload Summary File", type = ['xlsm'], key = "key2")

# second init
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'final_df' not in st.session_state:
    st.session_state.final_df = None
if 'formatted_files' not in st.session_state:
    st.session_state.formatted_files = []

# Show the list of uploaded files and corresponding business units
if st.session_state.uploaded_files:
    st.write("Uploaded Files and Corresponding Business Units:")
    for i, (file, bu) in enumerate(zip(st.session_state.uploaded_files, st.session_state.business_units)):
        st.write(f"{i+1}. {file.name} - {bu}")

# Provide an option to upload more files or process the files
if st.session_state.uploaded_files:
    if st.button("Process Files"):
        
        # Here, process all uploaded files together
        st.write("Processing the following files:")
        final_dfs = []
        st.session_state.formatted_files.clear()
        
        for file, bu, pswd, sht in zip(st.session_state.uploaded_files, st.session_state.business_units, st.session_state.passwords, st.session_state.sheets):
            
            # Call your read_workbook function to process each file
            st.write(f"Processing {file.name} for {bu}")
            df = read_workbook(bu, file, pswd, sht)
            df = format_columns(selected_month, bu, df)
            
            # Save the formatted DataFrame to a buffer
            output_buffer = io.BytesIO()
            df.to_excel(output_buffer, index=False, engine='openpyxl')
            output_buffer.seek(0)  # Move to the beginning of the buffer
            
            # Append the DataFrame buffer to formatted_files
            st.session_state.formatted_files.append((f"formatted_{file.name.replace('.xlsm', '.xlsx')}", output_buffer.read()))
            
            summary_df = df.groupby("department").agg(
                num_employees = ("employee_id", "nunique"),
                new_hire = ("is_new", "sum"),
                resign = ("is_resign", "sum"),
                total_basic_salary = ("basic_salary", "sum"),
                Bonus = ("bonus", "sum"),
                Overtime = ("overtime", "sum"),
                Allowance = ("allowance", "sum"),
                Incentives = ("incentives", "sum"),
                Claims = ("claims", "sum"),
                Others = ("others", "sum"),
                employer_statutory_payments = ("employer_statutory_payments", "sum"),
                total_staff_cost = ("total_staff_cost", "sum")
            ).reset_index()
            
            summary_df["Business Unit"] = bu
            final_dfs.append(summary_df)
            
        st.session_state.final_df = pd.concat(final_dfs, ignore_index = True)

        # rename and reorder columns
        column_renames = {
            "department": "Department",
            "num_employees": "Total HC",
            "new_hire": "New Join (HC)",
            "resign": "Resign (HC)",
            "total_basic_salary": "Basic Salary",
            "employer_statutory_payments": "Employer Statutory Payments",
            "total_staff_cost": "Total Staff Cost"
        }

        st.session_state.final_df = st.session_state.final_df.rename(columns = column_renames)

        # reordering cols
        last = st.session_state.final_df.columns[-1]
        new_column_order = [last] + [col for col in st.session_state.final_df.columns if col != last]
        st.session_state.final_df = st.session_state.final_df[new_column_order] 

        st.success("All files processed successfully!")
    
        # Create a zip file for formatted files
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            for formatted_filename, file_content in st.session_state.formatted_files:
                zip_file.writestr(formatted_filename, file_content)  # Write the content to the zip with correct extension

        zip_buffer.seek(0)  # Move the pointer to the beginning of the buffer

        # Provide a download button for the zip file
        st.download_button(
            label="Download All Formatted Files",
            data=zip_buffer,
            file_name="formatted_files.zip",
            mime="application/zip"
        )
        
        st.session_state.processed = True  # Set the processed state to True
        
if uploaded_xlsm_file is not None and st.session_state.final_df is not None:
    current_year = str(datetime.now().year)[2:]
    new_sheet_name = selected_month[:3] + "'" + current_year
                
    processed_file = add_df_to_xlsm(uploaded_xlsm_file, st.session_state.final_df, new_sheet_name)
                
    if processed_file:
        st.download_button(
            label = "Download Modified Excel Workbook",
            data = processed_file,
            file_name = "modified_workbook.xlsm",
            mime = "application/vnd.ms-excel.sheet.macroEnabled.12"
        )

# Reset the state to upload new files after processing
if st.button("Reset Uploads"):
    st.session_state.uploaded_files = []
    st.session_state.business_units = []
    st.success("Uploads reset. You can start uploading files again.")
